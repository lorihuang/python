import openpyxl

wb1 = openpyxl.load_workbook('test12_13.xlsx')
sheet1 = wb1.get_active_sheet()
wb2 = openpyxl.Workbook()
sheet2 = wb2.get_active_sheet()
col_num = int(sheet1.max_column)
row_num = int(sheet1.max_row)

for x in range(1,col_num+1):
    for y in range(1,row_num+1):
        sheet2.cell(row=x,column=y).value = sheet1.cell(row=y,column=x).value
        y += 1
    x += 1
wb2.save('test12_13_3.xlsx')
