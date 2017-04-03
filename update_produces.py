import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

PRICE_UPDATES = {'Garlic':3.07,
                 'Celery':1.19,
                 'Lemon':1.27}

number = 0
for row_num in range(2,sheet.max_row):
    produce_name = sheet.cell(row=row_num,column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num,column=2).value = PRICE_UPDATES[produce_name]
        number +=1
print('The work is done!\nYou have successful modifid ' + str(number) + ' recodes!')
wb.save('updated_produce_sales.xlsx')
