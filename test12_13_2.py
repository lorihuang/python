import openpyxl

x = int(input('Enter the first number: '))
y = int(input('Enter the second number: '))
wb1 = openpyxl.load_workbook('mul_table.xlsx')
sheet1 = wb1.get_active_sheet()
wb2 = openpyxl.Workbook()
sheet2 = wb2.get_active_sheet()
col_num = int(sheet1.max_column)
row_num = int(sheet1.max_row)

for i in range(1,x):
    for j in range(1,col_num+1):
        sheet2.cell(row=i,column=j).value = sheet1.cell(row=i,column=j).value
        j += 1
    i += 1
for i in range(x+y,row_num+y+1):
    for j in range(1,col_num+1):
        sheet2.cell(row=i,column=j).value = sheet1.cell(row=i-y,column=j).value
        j += 1
    i += 1
wb2.save('test12_13.xlsx')
