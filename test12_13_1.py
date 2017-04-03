import openpyxl
from openpyxl.styles import Font

x = int(input('Please input a number: '))
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
for t in range(1,x+1):
    sheet.cell(row=t+1,column=1).value = t
    sheet.cell(row=t+1,column=1).font = Font(b=True)
    sheet.cell(row=1,column=t+1).value = t
    sheet.cell(row=1,column=t+1).font = Font(b=True)
for y in range(2,x+2):
    for z in range(2,x+2):
        sheet.cell(row=y,column=z).value = (y-1)*(z-1)
        z += 1
    y += 1
wb.save('mul_table.xlsx')

